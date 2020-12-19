import pandas as pd
from os import listdir
import os
from os.path import isfile, join

INPUT_DIR = "inputfiles"
OUTPUT_DIR = "outputfiles"
SYMBOL_NAME = 'SILVER'
INSTRUMENT_TYPE = 'FUTCOM'
SHEET_NAME = 'At Contract Level'
# OUTPUT_FILE_PREFIX = 'output-'
OUTPUT_MERGED_FILE_NAME = 'output-merged-data.xlsx'

ENGINE = 'openpyxl'


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
        writeHeader = True
    else:
        writeHeader = False

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, header=writeHeader, **to_excel_kwargs)

    # save the workbook
    writer.save()

def isTransactionDoneForLeastMinExpiry(row, transaction_date, expiry_date_column, min_expiry_date,
                                       second_min_expiry_date):
    if row[transaction_date] <= min_expiry_date:
        expiry_date_to_check_against = min_expiry_date
    else:
        expiry_date_to_check_against = second_min_expiry_date

    return row[expiry_date_column] == expiry_date_to_check_against


def processExcelFile(inputdir, fileName):
    df = pd.read_excel(os.path.join(inputdir, fileName), sheet_name=SHEET_NAME, engine=ENGINE, header=None)

    df.columns = (df.loc[0:3].fillna('').apply(' '.join).str.strip())
    df = df.loc[4:].reset_index(drop=True)

    instrument_type_column = df.columns[4]
    symbol_column = df.columns[6]
    transaction_date_column = df.columns[2]
    expiry_date_column = df.columns[7]

    # print(df.shape)
    df.drop(df[(df[instrument_type_column] != INSTRUMENT_TYPE) | (df[symbol_column] != SYMBOL_NAME)].index, inplace=True)

    # print(df.shape)

    minExpiryDate = df[expiry_date_column].min()
    secondMinExpiryDate = (df[df[expiry_date_column] != minExpiryDate][expiry_date_column].min())

    # print("First Expiry: ", minExpiryDate)
    # print("Second Expiry: ", secondMinExpiryDate)

    # print(df.loc[:, [transaction_date_column, expiry_date_column]])

    for index, row in df.iterrows():
        if not isTransactionDoneForLeastMinExpiry(row, transaction_date_column, expiry_date_column, minExpiryDate,
                                                  secondMinExpiryDate):
            df.drop(index, inplace=True)

    # print(df.loc[:, [transaction_date_column, expiry_date_column]])

    output_file_path = os.path.join(OUTPUT_DIR, OUTPUT_MERGED_FILE_NAME)
    # df.to_excel(output_file_path, sheet_name=SHEET_NAME)
    append_df_to_excel(output_file_path, df, SHEET_NAME)
    print("Output written to: ", output_file_path)


def createOutputDir():
    if os.path.exists(OUTPUT_DIR):
        for root, dirs, files in os.walk(OUTPUT_DIR, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.removedirs(OUTPUT_DIR)
    os.mkdir(OUTPUT_DIR)

if __name__ == "__main__":
    createOutputDir()
    inputFiles = [f for f in listdir(INPUT_DIR) if isfile(join(INPUT_DIR, f))]
    inputFiles.sort()

    for inputFile in inputFiles:
        print ("Processing: " + os.path.join(INPUT_DIR, inputFile))
        processExcelFile(INPUT_DIR, inputFile)
